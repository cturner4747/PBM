
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import re

def generate_optumrx_excel(template_path, claims_df, output_path, ncpdp_override=None):
    wb = load_workbook(template_path)
    ws = wb.active
    start_row = 12

    def safe_set(ws, cell_ref, value):
        cell = ws[cell_ref]
        if not isinstance(cell, MergedCell):
            cell.value = value

    def normalize_ndc(ndc_raw):
        digits_only = re.sub(r"[^0-9]", "", str(ndc_raw))
        return digits_only.zfill(11)[:11] if digits_only else ""

    # Flexibly find the NDC column
    ndc_column = next((col for col in claims_df.columns if "ndc" in str(col).lower()), None)

    for idx, (_, row) in enumerate(claims_df.iterrows()):
        r = start_row + idx

        # Normalize PCN
        raw_pcn = str(row.get("PCN", "")).strip()
        if raw_pcn.endswith("00.00%"):
            pcn_val = raw_pcn.replace("00.00%", "")
        elif raw_pcn == "999900.00%":
            pcn_val = "9999"
        elif raw_pcn == "196000000.00%":
            pcn_val = "01960000"
        else:
            pcn_val = raw_pcn
        pcn_val = pcn_val.zfill(7)

        rx_val = str(row.get("Rx", "")).zfill(12)
        ndc_raw_val = row.get(ndc_column, "") if ndc_column else ""
        ndc_val = normalize_ndc(ndc_raw_val)

        safe_set(ws, f"A{r}", str(row.get("BIN", "")).zfill(6))
        safe_set(ws, f"B{r}", pcn_val)
        safe_set(ws, f"C{r}", "")
        safe_set(ws, f"D{r}", str(ncpdp_override).zfill(7) if ncpdp_override else "")
        safe_set(ws, f"E{r}", rx_val)
        safe_set(ws, f"F{r}", pd.to_datetime(row.get("Fill Date", "")).strftime("%m/%d/%Y") if pd.notna(row.get("Fill Date")) else "")
        safe_set(ws, f"G{r}", ndc_val)
        safe_set(ws, f"H{r}", "N")
        safe_set(ws, f"I{r}", "MAC Unit is below cost")
        safe_set(ws, f"J{r}", "")
        safe_set(ws, f"K{r}", row.get("Invoice Cost", ""))
        safe_set(ws, f"M{r}", row.get("TP Remit", ""))
        safe_set(ws, f"O{r}", row.get("Drug Name", ""))
        safe_set(ws, f"Q{r}", "")

    wb.save(output_path)
    return output_path
