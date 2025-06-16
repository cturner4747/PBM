
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

def generate_optumrx_excel(template_path, claims_df, output_path, ncpdp_override=None):
    """
    Populates the OptumRx MAC Appeal template with claim data, avoiding merged cells.
    """
    wb = load_workbook(template_path)
    ws = wb.active
    start_row = 12

    def safe_set(ws, cell_ref, value):
        cell = ws[cell_ref]
        if not isinstance(cell, MergedCell):
            cell.value = value

    for idx, (_, row) in enumerate(claims_df.iterrows()):
        r = start_row + idx
        safe_set(ws, f"A{r}", str(row.get("BIN", "")).zfill(6))
        safe_set(ws, f"B{r}", row.get("PCN", ""))
        safe_set(ws, f"C{r}", "")  # Carrier ID
        safe_set(ws, f"D{r}", str(ncpdp_override).zfill(7) if ncpdp_override else "")
        safe_set(ws, f"E{r}", str(row.get("Rx", "")))
        safe_set(ws, f"F{r}", pd.to_datetime(row.get("Fill Date", "")).strftime("%m/%d/%Y") if pd.notna(row.get("Fill Date")) else "")
        safe_set(ws, f"G{r}", str(row.get("Dispensed NDC", "")).replace("-", "").zfill(11)[:11])
        safe_set(ws, f"H{r}", "N")  # Compound
        safe_set(ws, f"I{r}", "MAC Unit is below cost")
        safe_set(ws, f"J{r}", "")
        safe_set(ws, f"K{r}", row.get("Invoice Cost", ""))
        safe_set(ws, f"M{r}", row.get("TP Remit", ""))
        safe_set(ws, f"O{r}", row.get("Drug Name", ""))
        safe_set(ws, f"Q{r}", "")

    wb.save(output_path)
    return output_path
